#https://blog.csdn.net/weixin_50702169/article/details/125127536 值得学习

import re
from openpyxl import load_workbook
from openpyxl.cell import MergedCell

class to_html():
    def __init__(self, file, save_file, sheet_name):
        self.file = file      # 文件路径
        self.save_file = save_file        # HTML保存路径
        self.sheet_name = sheet_name      # Sheet名

    def creat_html(self):
        wb = load_workbook(filename=self.file)
        sheet = wb[self.sheet_name]
        cell_dic = {}        # 用于储存所有合并单元格的左上单元格对象
        col_width = {}       # 用于储存所有列的列宽, px
        row_height = {}      # 用于储存所有列的行高, px

        # 查询列宽
        for col in sheet.columns:
            pat = r"[A-Z]+"
            pat = re.compile(pat)
            colname = pat.findall(col[0].coordinate)[0]       # 分离字母和数字，取出列字母名称
            px = round(sheet.column_dimensions[colname].width * 10)  # 读出列宽换算为像素
            col_width[colname] = px

        # 查询行高
        for row in sheet.rows:
            pat = r"[A-Z]+(\d+)"
            pat = re.compile(pat)
            rowid = int(pat.findall(row[0].coordinate)[0])       # 分离字母和数字，取出行数字序号
            px = sheet.row_dimensions[rowid].height        # 读出行高换算为像素
            if px is None:
                px = 13.5
            row_height[str(rowid)] = px

        # 找出所有合并区域的行高，列宽，向右合并距离，向下合并距离
        for merged_range in sheet.merged_cells.ranges:
            now_width = 0   # 定义列宽
            now_height = 0    # 定义行高
            for i in range(merged_range.min_col, merged_range.max_col + 1):
                coord = sheet.cell(row=1, column=i).coordinate      # 位置标识，例如：A1
                pat = r"[A-Z]+"
                pat = re.compile(pat)
                colname = pat.findall(coord)[0]  # 分离字母和数字，取出列字母名称
                now_width = now_width + col_width[colname]
            for i in range(merged_range.min_row, merged_range.max_row + 1):
                coord = sheet.cell(row=i, column=1).coordinate  # 位置标识，例如：A1
                pat = r"[A-Z]+(\d+)"
                pat = re.compile(pat)
                colindex = pat.findall(coord)[0]  # 分离字母和数字，取出列数字名称
                now_height = now_height + row_height[colindex]

            now_width = int(now_width)        # 合并单元格列宽（所有子单元格相加）
            now_height = int(now_height)      # 合并单元格行高（所有子单元格相加）

            cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)  # 选择合并区域左上单元格
            colspan = merged_range.max_col - merged_range.min_col + 1  # 向右合并长度
            rowspan = merged_range.max_row - merged_range.min_row + 1  # 向下合并长度
            cell_dic[cell] = (now_height, now_width, colspan, rowspan)

        html = '''<table border="1">'''
        rows_list = list(sheet.rows)
        row_count = len(rows_list)
        processed_rows = 0

        # 如果行数小于5行，则全部都处理
        if row_count <= 5:
            processed_rows = row_count

        # 开始写入数据到table标签
        for row in sheet.rows:
            tr = '''<tr>'''
            for cell in row:
                td = ""
                if cell in cell_dic:       # 判断是否为合并单元格左上单元格
                    if cell.value is None:
                        text = ''
                    else:
                        text = cell.value

                    if cell.alignment.vertical is not None:
                        vertical = f'''vertical-align: {cell.alignment.vertical};'''  # 水平位置
                    else:
                        vertical = ''
                    if cell.alignment.horizontal is not None:
                        horizontal = f'''text-align: {cell.alignment.horizontal};'''  # 垂直位置
                    else:
                        horizontal = ''

                    font_size = str(int(cell.font.size) + 3)  # 字体大小
                    font_weight = '700' if cell.font.b else '400'  # 字体是否加粗
                    style = f'''"font-family: Arial, Helvetica, sans-serif; font-size: {font_size}px; font-weight: {font_weight}; font-style: normal;{vertical}{horizontal}"'''
                    td = f'''<td height="{cell_dic[cell][0]}" width="{cell_dic[cell][1]}" colspan="{cell_dic[cell][2]}" rowspan="{cell_dic[cell][3]}" style={style}>{text}</td>'''
                else:
                    if not isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
                        if cell.alignment.vertical is not None:
                            vertical = f'''vertical-align: {cell.alignment.vertical};'''  # 水平位置
                        else:
                            vertical = ''
                        if cell.alignment.horizontal is not None:
                            horizontal = f'''text-align: {cell.alignment.horizontal};'''  # 垂直位置
                        else:
                            horizontal = ''
                        pat = r"([A-Z]+)(\d+)"
                        pat = re.compile(pat)
                        cell_name = pat.findall(cell.coordinate)[0][0]
                        cell_index = pat.findall(cell.coordinate)[0][1]
                        font_size = str(int(cell.font.size) + 3)  # 字体大小
                        font_weight = '700' if cell.font.b else '400'  # 字体是否加粗
                        style = f'''"font-family: Arial, Helvetica, sans-serif; font-size: {font_size}px; font-weight: {font_weight}; font-style: normal;{vertical}{horizontal}"'''
                        if cell.value is not None:
                            td = f'''<td height="{row_height[cell_index]}" width="{col_width[cell_name]}" style={style} >{cell.value}</td>'''
                        else:
                            td = f'''<td height="{row_height[cell_index]}" width="{col_width[cell_name]}"></td>'''

                tr = tr + td
            tr = tr + '''</tr>'''

            # 仅处理第1、2行和倒数1-5行的数据
            if processed_rows < 2 or (row_count <= 5 or processed_rows >= (row_count - 5)):
                html = html + tr

            processed_rows += 1

        html = html + '''</table>'''
        
        with open(self.save_file,'w',encoding='utf-8') as f:
                    f.write(html)

        return html
    
# if __name__=='__main__':
#     data=to_html("/opt/performance/performance_tpcc_everyday/html/tpcc/performance_tpcc_all.xlsx","/opt/performance/performance_tpcc_everyday/html/tpcc/performance_tpcc_all.html","summary")
#     data.creat_html()
