
import os
import re
from lib.tpQconfig import *
from lib.Logger import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class ResultProcess:
    def __init__(self):
        self.tpcc_run_result_directory = Userinof.tpcc_run_result_directory.format(Userinof.version_date)
        self.date = Userinof.version_date
        self.tpcc_excel_path = os.path.join(os.getcwd(),'html','tpcc')
        self.tpcc_moban = os.path.join(os.getcwd(),'doc','performance_tpcc_moban.xlsx')
        self.tpcc_all_moban=os.path.join(os.getcwd(),'doc','performance_tpcc_all_moban.xlsx')
        self.tpcc_run_threads = Userinof.tpcc_run_threads
        self.tpcc_run_number = Userinof.tpcc_run_number

    def read_tpcc_log_list(self):

        filename=os.path.join(self.tpcc_run_result_directory,'log_tpcc_list')
        
        with open(filename,'r+') as f:
            content = f.read().split()

        return content 
    
    def tpcc_resulte(self,log_file):
        
        tpcc_terminals=''
        tpcc_Session_Start=''
        tpcc_Session_End=''
        tpcc_tpmC=''
        tpcc_tpmtotal=''
        tpcc_Transaction_Count=''

        with open(log_file, 'r', encoding="utf-8") as f:
            all_content = f.read()

        tpcc_terminals = re.findall(r'terminals=(.*?)\n',all_content,re.S)
        tpcc_Session_Start = re.findall(r'Session Start.*?= (.*?)\n',all_content,re.S)
        tpcc_Session_End = re.findall(r'Session End.*?= (.*?)\n',all_content,re.S)
        tpcc_tpmC = re.findall(r'Measured tpmC.*?= (.*?)\n',all_content,re.S)
        tpcc_tpmtotal = re.findall(r'Measured tpmTOTAL = (.*?)\n',all_content,re.S)
        tpcc_Transaction_Count = re.findall(r'Transaction Count = (.*?)\n',all_content,re.S)

        # print(tpcc_terminals,tpcc_Session_Start,tpcc_Session_End,tpcc_tpmC,tpcc_tpmtotal,tpcc_Transaction_Count)
        return tpcc_terminals,tpcc_Session_Start,tpcc_Session_End,tpcc_tpmC,tpcc_tpmtotal,tpcc_Transaction_Count
    

    def performance_tpcc_excel(self, log_files):
        logger.info("performance_tpcc_excel")
        date_folder = os.path.join(self.tpcc_excel_path, self.date)
        os.makedirs(date_folder, exist_ok=True)

        excel_file = os.path.join(date_folder, 'performance_tpcc_{}.xlsx'.format(self.date))

        shutil.copyfile(self.tpcc_moban, excel_file)

        wb_all = load_workbook(excel_file)
        ws_today = wb_all["today"]

        for log_file in log_files:
            tpcc_terminals,tpcc_Session_Start,tpcc_Session_End,tpcc_tpmC,tpcc_tpmtotal,tpcc_Transaction_Count = self.tpcc_resulte(log_file)
            # print(tpcc_terminals,tpcc_Session_Start,tpcc_Session_End,tpcc_tpmC,tpcc_tpmtotal,tpcc_Transaction_Count)
            #today
            # 获取"today"表格的下一个空行
            next_row = ws_today.max_row + 1

            # date
            ws_today.cell(row=next_row, column=1, value=self.date)
            # tpcc_terminals
            ws_today.cell(row=next_row, column=2, value=tpcc_terminals[0] if tpcc_terminals else "")
            # tpcc_Session_Start
            ws_today.cell(row=next_row, column=3, value=tpcc_Session_Start[0] if tpcc_Session_Start else "")
            # tpcc_Session_End
            ws_today.cell(row=next_row, column=4, value=tpcc_Session_End[0] if tpcc_Session_End else "")
            # tpmC
            ws_today.cell(row=next_row, column=5, value=float(tpcc_tpmC[0]) if tpcc_tpmC else "")
            # tpmtotal
            ws_today.cell(row=next_row, column=6, value=float(tpcc_tpmtotal[0]) if tpcc_tpmtotal else "")
            # Transaction Count
            ws_today.cell(row=next_row, column=7, value=float(tpcc_Transaction_Count[0]) if tpcc_Transaction_Count else "")


            # Set Border Style
            border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws_today.iter_rows(min_row=next_row-1, max_row=next_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = border_style

        ws_today.column_dimensions[get_column_letter(3)].width = 20
        ws_today.column_dimensions[get_column_letter(4)].width = 20
        ws_today.column_dimensions[get_column_letter(5)].width = 12
        ws_today.column_dimensions[get_column_letter(6)].width = 12
        
        wb_all.save(excel_file)

        return excel_file
    
    def performance_tpcc_excel_all(self,start_test_time,excel_file):
        logger.info("performance_tpcc_excel_all")
        excel_file_all = os.path.join(self.tpcc_excel_path, "performance_tpcc_all.xlsx")

        if not os.path.exists(excel_file_all):
            shutil.copyfile(self.tpcc_all_moban, excel_file_all)

        wb = load_workbook(excel_file) 
        wb_today = wb["today"]
        
        wb_all = load_workbook(excel_file_all)
        

        #summary
        wb_all_summary = wb_all["summary"]
        # thread_nums = [int(x) for x in self.tpcc_run_threads.split()]
        thread_nums = re.findall(r'\d+', self.tpcc_run_threads)
        wb_all_summary.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3+len(thread_nums))
        for i, thread_num in enumerate(thread_nums, start=4):
            wb_all_summary.cell(row=2, column=i, value=thread_num)
            # 设置列宽为12和居中
            column_letter = get_column_letter(i)
            wb_all_summary.column_dimensions[column_letter].width = 12.5
            cell = wb_all_summary.cell(row=2, column=i, value=thread_num)
            alignment = Alignment(horizontal="center", vertical="center")
            cell.alignment = alignment

         # last_row
        last_row = len(list(wb_all_summary.iter_rows())) + 1

        wb_all_summary.cell(row=last_row, column=1, value=self.date)
        wb_all_summary.cell(row=last_row, column=2, value=start_test_time)
        wb_all_summary.cell(row=last_row, column=3, value=self.tpcc_run_number)

        # 计算并填充tpmc和tpmtotal
        # thread_nums = re.findall(r'\d+', self.tpcc_run_threads)
        for i, thread_num in enumerate(thread_nums, start=4):
            # tpmc_sum = sum([float(cell.value) for row in wb_today.iter_rows(min_row=2) for idx, cell in enumerate(row) if idx == 1 and cell.value == thread_num])
            tpmc_sum = 0.0
            # 迭代wb_today中的每一行，从第二行开始
            for row in wb_today.iter_rows(min_row=2):
                for idx,cell in enumerate(row):
                    if idx == 1 and cell.value == thread_num:
                        tpmc_sum += float(row[4].value)# 获取第五列（索引为4）的值
            tpmc_avg = tpmc_sum / float(self.tpcc_run_number)
            wb_all_summary.cell(row=last_row, column=i, value='{:.2f}'.format(tpmc_avg))


        # 遍历所有边框
        for row_idx in range(1, last_row+1):
            for col_idx in range(1, len(thread_nums) + 4):  
                cell = wb_all_summary.cell(row=row_idx, column=col_idx)
                
                # 设置所有边的边框样式
                border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.border = border_style

        
        #today
        wb_all_today = wb_all["today"]
        # 复制数据从wb_today到wb_all_today
        for row_idx, row in enumerate(wb_today.iter_rows(min_row=1), start=1):
            for col_idx, cell in enumerate(row, start=1):
                wb_all_today.cell(row=row_idx, column=col_idx, value=cell.value)
                
            # Set Border Style
            border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in wb_all_today.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=7):
                for cell in row:
                    cell.border = border_style


        #all
        wb_all_all = wb_all["all"]
        last_row = len(list(wb_all_all.iter_rows()))+1
        for row_idx, row in enumerate(wb_today.iter_rows(min_row=2), start=last_row):  # 从wb_today的第二行开始
            for col_idx, cell in enumerate(row, start=1):
                wb_all_all.cell(row=row_idx, column=col_idx, value=cell.value)

            border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in wb_all_all.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=7):
                for cell in row:
                    cell.border = border_style

        wb_all.save(excel_file_all)

        return excel_file_all